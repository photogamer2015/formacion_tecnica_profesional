"""
Vistas adicionales para Pagos, Historial de matriculados y Estudiantes.

Diseño:
- Todas usan el decorador @matricula_requerida (admin + asesor pueden ver).
- Las exportaciones a Excel usan openpyxl y devuelven un HttpResponse con el archivo.
- Filtros por GET querystring (q, curso, modalidad, estado, año, mes).
"""

from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Count, Prefetch, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST

from .forms import AbonoForm, RecuperacionPendienteForm
from .models import Abono, Curso, Estudiante, Matricula, RecuperacionPendiente
from .permisos import matricula_requerida


# ═════════════════════════════════════════════════════════════════
# Constantes
# ═════════════════════════════════════════════════════════════════

MESES_ES = [
    '', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre',
]


# ═════════════════════════════════════════════════════════════════
# Helpers de Excel
# ═════════════════════════════════════════════════════════════════

def _build_excel_response(filename, sheet_name, headers, rows, totals=None):
    """
    Genera un .xlsx en memoria y lo devuelve como HttpResponse para descarga.

    headers: lista de strings (encabezados de columna)
    rows: lista de listas (datos)
    totals: dict opcional {col_idx_0based: total} para fila final
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.page import PageMargins

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel limita a 31 chars

    # ── Estilos ──
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='1A237E')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )
    total_font = Font(bold=True, color='1A237E', size=11)
    total_fill = PatternFill('solid', fgColor='FFF8E1')
    data_align_wrap = Alignment(vertical='center', wrap_text=True)

    # ── Título de la hoja en fila 1 ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=sheet_name)
    title_cell.font = Font(bold=True, size=14, color='1A237E')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    # ── Encabezados en fila 2 ──
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[2].height = 30

    # ── Datos a partir de fila 3 ──
    for row_idx, row_data in enumerate(rows, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = data_align_wrap

    # ── Fila de totales ──
    if totals:
        total_row_idx = len(rows) + 3
        # Etiqueta "TOTAL" en la primera columna
        cell = ws.cell(row=total_row_idx, column=1, value='TOTAL')
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='right', vertical='center')

        for col_idx_0, value in totals.items():
            cell = ws.cell(row=total_row_idx, column=col_idx_0 + 1, value=value)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    # ── Auto-ancho aproximado por columna (con tope y mínimo razonables) ──
    for col_idx, header in enumerate(headers, start=1):
        max_length = len(str(header))
        for row_data in rows:
            if col_idx - 1 < len(row_data):
                value = row_data[col_idx - 1]
                if value is not None:
                    # Considera la línea más larga si hay saltos de línea
                    longest = max((len(s) for s in str(value).split('\n')), default=0)
                    max_length = max(max_length, longest)
        # Un poco más amplio para que no se corte el texto al imprimir
        width = min(max(max_length + 3, 10), 38)
        ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = width

    # ── Congelar encabezados ──
    ws.freeze_panes = 'A3'

    # ── Configuración de impresión profesional ──
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A3 if len(headers) > 10 else ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # Tantas páginas como sean necesarias en alto
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.4,
                                   header=0.2, footer=0.2)
    # Repetir las dos primeras filas (título + encabezado) en cada página al imprimir
    ws.print_title_rows = '1:2'
    # Encabezado y pie de página al imprimir
    ws.oddHeader.center.text = f"&B{sheet_name}"
    ws.oddHeader.center.size = 11
    ws.oddHeader.center.color = "1A237E"
    ws.oddFooter.left.text = "Formación Profesional EC"
    ws.oddFooter.right.text = "Página &P de &N"
    ws.oddFooter.left.size = 9
    ws.oddFooter.right.size = 9

    # ── Devolver como HttpResponse ──
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _filtrar_matriculas(request):
    """
    Aplica filtros comunes a una queryset de Matricula según los GET params.
    Devuelve (queryset filtrada, dict de filtros aplicados).
    """
    qs = Matricula.objects.select_related(
        'estudiante', 'curso', 'curso__categoria', 'jornada', 'registrado_por'
    )

    estado = request.GET.get('estado', '').strip()
    curso_id = request.GET.get('curso', '').strip()
    modalidad = request.GET.get('modalidad', '').strip()
    anio = request.GET.get('anio', '').strip()
    mes = request.GET.get('mes', '').strip()
    q = request.GET.get('q', '').strip()

    if curso_id:
        qs = qs.filter(curso_id=curso_id)
    if modalidad in ('presencial', 'online'):
        qs = qs.filter(modalidad=modalidad)
    if anio.isdigit():
        qs = qs.filter(fecha_matricula__year=int(anio))
    if mes.isdigit() and 1 <= int(mes) <= 12:
        qs = qs.filter(fecha_matricula__month=int(mes))
    if q:
        qs = qs.filter(
            Q(estudiante__cedula__icontains=q)
            | Q(estudiante__apellidos__icontains=q)
            | Q(estudiante__nombres__icontains=q)
            | Q(curso__nombre__icontains=q)
        )

    # Filtro por estado (Pagado/Parcial/Pendiente/Retiro) — se hace en Python
    # porque saldo es una @property, no un campo de DB.
    # Para mantener qs como queryset, traduzco el estado a condiciones:
    if estado == 'Retiro':
        qs = qs.filter(estado='retiro_voluntario')
    elif estado == 'Pagado':
        qs = qs.filter(valor_pagado__gte=models_F('valor_curso')).exclude(estado='retiro_voluntario')
    elif estado == 'Parcial':
        qs = qs.filter(valor_pagado__gt=0, valor_pagado__lt=models_F('valor_curso')).exclude(estado='retiro_voluntario')
    elif estado == 'Pendiente':
        qs = qs.filter(Q(valor_pagado=0) | Q(valor_pagado__isnull=True)).exclude(estado='retiro_voluntario')

    return qs, {
        'estado': estado,
        'curso': curso_id,
        'modalidad': modalidad,
        'anio': anio,
        'mes': mes,
        'q': q,
    }


def _resumen_abonos(abonos):
    """
    Agrupa los pagos de una matrícula por tipo de pago y por método.
    Devuelve estructuras simples para poder pintarlas en tablas y exportarlas.
    """
    tipos = {}
    metodos = {}

    for abono in abonos:
        monto = abono.monto or Decimal('0.00')

        tipo_label = abono.get_tipo_pago_display()
        if not abono.cuenta_para_saldo and abono.tipo_pago == 'recuperacion':
            tipo_label = f'{tipo_label} (aparte)'
        tipo = tipos.setdefault(tipo_label, {'label': tipo_label, 'total': Decimal('0.00'), 'count': 0})
        tipo['total'] += monto
        tipo['count'] += 1

        metodo_label = abono.get_metodo_display()
        if abono.metodo in ('transferencia', 'tarjeta') and abono.banco:
            metodo_label = f'{metodo_label} · {abono.get_banco_display()}'
        metodo = metodos.setdefault(metodo_label, {'label': metodo_label, 'total': Decimal('0.00'), 'count': 0})
        metodo['total'] += monto
        metodo['count'] += 1

    return {
        'tipos': list(tipos.values()),
        'metodos': list(metodos.values()),
        'total_movimientos': sum((x['count'] for x in tipos.values()), 0),
    }


def _adjuntar_resumen_abonos(matriculas):
    """Agrega a cada matrícula el resumen de sus abonos ya prefetched."""
    for matricula in matriculas:
        abonos = getattr(matricula, 'abonos_para_resumen', None)
        if abonos is None:
            abonos = list(matricula.abonos.all())
        matricula.resumen_abonos = _resumen_abonos(abonos)
    return matriculas


# Importación tardía para evitar circular imports en algunos casos
from django.db.models import F as models_F


# ═════════════════════════════════════════════════════════════════
# Pagos
# ═════════════════════════════════════════════════════════════════

@matricula_requerida
def pagos_lista(request):
    """
    Vista centrada en lo financiero. Permite filtrar matrículas por:
    - Estado de pago: Pagado, Parcial, Pendiente
    - Curso (buscador)
    - Modalidad
    """
    qs, filtros = _filtrar_matriculas(request)
    qs = qs.prefetch_related(
        Prefetch('abonos', queryset=Abono.objects.order_by('fecha', 'id'), to_attr='abonos_para_resumen')
    ).order_by('-fecha_matricula', '-id')
    matriculas = _adjuntar_resumen_abonos(list(qs))

    # Estadísticas globales (con los filtros aplicados, excepto el de estado)
    qs_sin_estado = Matricula.objects.select_related('curso').all()
    if filtros['curso']:
        qs_sin_estado = qs_sin_estado.filter(curso_id=filtros['curso'])
    if filtros['modalidad']:
        qs_sin_estado = qs_sin_estado.filter(modalidad=filtros['modalidad'])
    if filtros['anio'].isdigit():
        qs_sin_estado = qs_sin_estado.filter(fecha_matricula__year=int(filtros['anio']))

    totales = {
        'total_matriculas': qs_sin_estado.count(),
        'total_facturado': qs_sin_estado.aggregate(s=Sum('valor_curso'))['s'] or Decimal('0.00'),
        'total_cobrado': qs_sin_estado.aggregate(s=Sum('valor_pagado'))['s'] or Decimal('0.00'),
        'total_pendiente': Decimal('0.00'),
        'total_retiro': Decimal('0.00'),
    }

    # Conteo por estado y recálculo de saldo pendiente sin retiros
    todos_los_pagos = list(qs_sin_estado.values('valor_curso', 'valor_pagado', 'estado'))
    conteo_estado = {'Pagado': 0, 'Parcial': 0, 'Pendiente': 0, 'Retiro': 0}
    
    for p in todos_los_pagos:
        vc = p['valor_curso'] or Decimal('0.00')
        vp = p['valor_pagado'] or Decimal('0.00')
        st = p['estado']
        
        if st == 'retiro_voluntario':
            conteo_estado['Retiro'] += 1
            totales['total_retiro'] += (vc - vp)
        else:
            totales['total_pendiente'] += (vc - vp)
            if vp >= vc and vc > 0:
                conteo_estado['Pagado'] += 1
            elif vp > 0:
                conteo_estado['Parcial'] += 1
            else:
                conteo_estado['Pendiente'] += 1

    cursos = Curso.objects.filter(activo=True).order_by('nombre')
    anios = sorted(
        set(Matricula.objects.dates('fecha_matricula', 'year').values_list('fecha_matricula__year', flat=True)),
        reverse=True
    )

    return render(request, 'pagos/lista.html', {
        'matriculas': matriculas,
        'cursos': cursos,
        'anios': anios,
        'filtros': filtros,
        'totales': totales,
        'conteo_estado': conteo_estado,
    })


@matricula_requerida
def pagos_export(request):
    """Descarga los pagos filtrados como Excel."""
    qs, filtros = _filtrar_matriculas(request)
    qs = qs.prefetch_related(
        Prefetch('abonos', queryset=Abono.objects.order_by('fecha', 'id'), to_attr='abonos_para_resumen')
    ).order_by('-fecha_matricula', '-id')

    headers = [
        'Fecha matrícula', 'Cédula', 'Apellidos', 'Nombres',
        'Curso', 'Categoría', 'Modalidad', 'Sede / Plataforma',
        'Jornada', 'Día (inicio jornada)', 'Horario',
        'Valor curso', 'Valor pagado', 'Saldo', 'Estado',
        'Tipos de pago', 'Métodos de pago',
        'Asistencia',
    ]

    rows = []
    total_curso = Decimal('0.00')
    total_pagado = Decimal('0.00')
    total_saldo = Decimal('0.00')

    for m in _adjuntar_resumen_abonos(list(qs)):
        tipos_pago = '; '.join(
            f"{x['label']}: ${x['total']:.2f} ({x['count']})"
            for x in m.resumen_abonos['tipos']
        ) or 'Sin pagos'
        metodos_pago = '; '.join(
            f"{x['label']}: ${x['total']:.2f} ({x['count']})"
            for x in m.resumen_abonos['metodos']
        ) or 'Sin pagos'

        # ── Datos de jornada ──
        if m.jornada:
            jornada_txt = m.jornada.descripcion_legible
            dia_inicio = m.jornada.fecha_inicio.strftime('%d/%m/%Y') if m.jornada.fecha_inicio else '—'
            if m.jornada.hora_inicio and m.jornada.hora_fin:
                horario_txt = f"{m.jornada.hora_inicio.strftime('%H:%M')} – {m.jornada.hora_fin.strftime('%H:%M')}"
            else:
                horario_txt = '—'
        else:
            jornada_txt = '—'
            dia_inicio = '—'
            horario_txt = '—'

        rows.append([
            m.fecha_matricula.strftime('%d/%m/%Y') if m.fecha_matricula else '',
            m.estudiante.cedula,
            m.estudiante.apellidos,
            m.estudiante.nombres,
            m.curso.nombre,
            m.curso.categoria.nombre if m.curso.categoria else '—',
            m.get_modalidad_display(),
            m.sede,
            jornada_txt,
            dia_inicio,
            horario_txt,
            float(m.valor_curso or 0),
            float(m.valor_pagado or 0),
            float(m.saldo or 0),
            m.estado_pago,
            tipos_pago,
            metodos_pago,
            '',  # Asistencia: columna en blanco para firma
        ])
        total_curso += m.valor_curso or Decimal('0.00')
        total_pagado += m.valor_pagado or Decimal('0.00')
        total_saldo += m.saldo or Decimal('0.00')

    totals = {
        11: float(total_curso),
        12: float(total_pagado),
        13: float(total_saldo),
    }

    fecha_str = datetime.now().strftime('%Y%m%d_%H%M')
    sufijo = ''
    if filtros['estado']:
        sufijo += f"_{filtros['estado'].lower()}"
    if filtros['anio']:
        sufijo += f"_{filtros['anio']}"
    filename = f'pagos{sufijo}_{fecha_str}.xlsx'

    return _build_excel_response(
        filename=filename,
        sheet_name='Reporte de Pagos',
        headers=headers,
        rows=rows,
        totals=totals,
    )


@matricula_requerida
def pagos_export_pdf(request):
    """Descarga los pagos filtrados como PDF horizontal con columna de asistencia."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, A3
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    except ImportError:
        return HttpResponse(
            'Para exportar a PDF instala reportlab: pip install reportlab',
            status=500, content_type='text/plain; charset=utf-8',
        )

    qs, filtros = _filtrar_matriculas(request)
    qs = qs.prefetch_related(
        Prefetch('abonos', queryset=Abono.objects.order_by('fecha', 'id'), to_attr='abonos_para_resumen')
    ).order_by('-fecha_matricula', '-id')
    matriculas = _adjuntar_resumen_abonos(list(qs))

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A3),
        leftMargin=0.7*cm, rightMargin=0.7*cm, topMargin=1.0*cm, bottomMargin=0.8*cm,
        title='Estado de Pagos',
    )
    styles = getSampleStyleSheet()
    titulo_st = ParagraphStyle(
        'titulo_pagos', parent=styles['Title'],
        textColor=colors.HexColor('#1A237E'), fontSize=15,
        alignment=1, spaceAfter=4,
    )
    sub_st = ParagraphStyle(
        'sub_pagos', parent=styles['Normal'],
        textColor=colors.HexColor('#666666'), fontSize=9,
        alignment=1, spaceAfter=10,
    )
    cell_st = ParagraphStyle(
        'cell', parent=styles['Normal'],
        fontName='Helvetica', fontSize=7, leading=8.5,
    )
    cell_bold_st = ParagraphStyle(
        'cell_bold', parent=cell_st, fontName='Helvetica-Bold',
    )

    elementos = [
        Paragraph('Estado de Pagos — Formación Profesional EC', titulo_st),
        Paragraph(
            f'Generado el {date.today().strftime("%d/%m/%Y")} · '
            f'{len(matriculas)} registro(s)',
            sub_st,
        ),
    ]

    headers = [
        'Fecha', 'Cédula', 'Estudiante', 'Curso', 'Modalidad',
        'Jornada', 'Día', 'Valor', 'Pagado', 'Tipo de pago',
        'Método', 'Saldo', 'Estado', 'Asistencia',
    ]
    # Header como Paragraphs (los wraps se hacen automáticamente)
    header_st = ParagraphStyle(
        'header_st', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=8,
        textColor=colors.whitesmoke, alignment=1, leading=9,
    )
    data = [[Paragraph(h, header_st) for h in headers]]
    total_curso = total_pagado = total_saldo = 0.0

    for m in matriculas:
        if m.jornada:
            jornada_txt = m.jornada.descripcion_legible or '—'
            dia_inicio = m.jornada.fecha_inicio.strftime('%d/%m/%Y') if m.jornada.fecha_inicio else '—'
        else:
            jornada_txt = '—'
            dia_inicio = '—'

        tipos_pago = '<br/>'.join(
            f"{x['label']}: ${x['total']:.2f}"
            for x in m.resumen_abonos['tipos']
        ) or 'Sin pagos'
        metodos_pago = '<br/>'.join(
            f"{x['label']}: ${x['total']:.2f}"
            for x in m.resumen_abonos['metodos']
        ) or 'Sin método'

        estado_txt = m.estado_pago or '—'

        data.append([
            Paragraph(m.fecha_matricula.strftime('%d/%m/%Y') if m.fecha_matricula else '', cell_st),
            Paragraph(m.estudiante.cedula or '', cell_bold_st),
            Paragraph(f'{m.estudiante.apellidos} {m.estudiante.nombres}'.strip(), cell_st),
            Paragraph(m.curso.nombre or '', cell_st),
            Paragraph(m.get_modalidad_display() or '', cell_st),
            Paragraph(jornada_txt, cell_st),
            Paragraph(dia_inicio, cell_st),
            Paragraph(f'${float(m.valor_curso or 0):.2f}', cell_st),
            Paragraph(f'<font color="#2e7d32"><b>${float(m.valor_pagado or 0):.2f}</b></font>', cell_st),
            Paragraph(tipos_pago, cell_st),
            Paragraph(metodos_pago, cell_st),
            Paragraph(f'<font color="{"#c62828" if (m.saldo or 0) > 0 else "#2e7d32"}"><b>${float(m.saldo or 0):.2f}</b></font>', cell_st),
            Paragraph(estado_txt, cell_st),
            '',  # Asistencia (firma) — vacío para llenar a mano
        ])
        total_curso += float(m.valor_curso or 0)
        total_pagado += float(m.valor_pagado or 0)
        total_saldo += float(m.saldo or 0)

    # Fila de totales
    data.append([
        Paragraph('', cell_st), Paragraph('', cell_st), Paragraph('', cell_st),
        Paragraph('', cell_st), Paragraph('', cell_st), Paragraph('', cell_st),
        Paragraph('<b>TOTAL</b>', cell_bold_st),
        Paragraph(f'<b>${total_curso:.2f}</b>', cell_bold_st),
        Paragraph(f'<b>${total_pagado:.2f}</b>', cell_bold_st),
        Paragraph('', cell_st), Paragraph('', cell_st),
        Paragraph(f'<b>${total_saldo:.2f}</b>', cell_bold_st),
        Paragraph('', cell_st), '',
    ])

    # ── Anchos de columna explícitos para A3 horizontal (≈ 41 cm de ancho útil) ──
    # Suman ~39.5 cm dejando margen
    col_widths = [
        1.7*cm,  # Fecha
        2.0*cm,  # Cédula
        4.5*cm,  # Estudiante
        3.5*cm,  # Curso
        1.8*cm,  # Modalidad
        3.5*cm,  # Jornada
        1.8*cm,  # Día
        1.8*cm,  # Valor
        1.8*cm,  # Pagado
        4.8*cm,  # Tipo de pago
        4.8*cm,  # Método
        1.8*cm,  # Saldo
        2.0*cm,  # Estado
        3.7*cm,  # Asistencia
    ]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A237E')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 1), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#BBBBBB')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F8F9FB')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFF8E1')),
        # Línea de firma en columna Asistencia
        ('LINEBELOW', (-1, 1), (-1, -2), 0.5, colors.HexColor('#888888')),
    ]))
    elementos.append(table)
    doc.build(elementos)

    pdf_bytes = buf.getvalue()
    buf.close()
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    fecha_str = date.today().strftime('%Y%m%d')
    sufijo = ''
    if filtros['estado']:
        sufijo += f"_{filtros['estado'].lower()}"
    filename = f'pagos{sufijo}_{fecha_str}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ═════════════════════════════════════════════════════════════════
# Historial de matriculados (por año / mes)
# ═════════════════════════════════════════════════════════════════

@matricula_requerida
def historial_lista(request):
    """
    Historial de matrículas agrupado por año y mes.
    Permite filtrar por año, mes, curso y modalidad.
    """
    qs, filtros = _filtrar_matriculas(request)
    qs = qs.order_by('-fecha_matricula', '-id')

    # Agrupar por año → mes → matrículas
    grupos = defaultdict(lambda: defaultdict(list))
    totales_por_anio = defaultdict(lambda: {'count': 0, 'facturado': Decimal('0.00'), 'cobrado': Decimal('0.00')})
    totales_por_mes = defaultdict(lambda: {'count': 0, 'facturado': Decimal('0.00'), 'cobrado': Decimal('0.00')})

    for m in qs:
        anio = m.fecha_matricula.year
        mes = m.fecha_matricula.month
        grupos[anio][mes].append(m)

        totales_por_anio[anio]['count'] += 1
        totales_por_anio[anio]['facturado'] += m.valor_curso or Decimal('0.00')
        totales_por_anio[anio]['cobrado'] += m.valor_pagado or Decimal('0.00')

        key = (anio, mes)
        totales_por_mes[key]['count'] += 1
        totales_por_mes[key]['facturado'] += m.valor_curso or Decimal('0.00')
        totales_por_mes[key]['cobrado'] += m.valor_pagado or Decimal('0.00')

    # Convertir a lista ordenada para el template
    estructura = []
    for anio in sorted(grupos.keys(), reverse=True):
        meses_dict = grupos[anio]
        meses_lista = []
        for mes in sorted(meses_dict.keys(), reverse=True):
            meses_lista.append({
                'numero': mes,
                'nombre': MESES_ES[mes],
                'matriculas': meses_dict[mes],
                'totales': totales_por_mes[(anio, mes)],
            })
        estructura.append({
            'anio': anio,
            'meses': meses_lista,
            'totales': totales_por_anio[anio],
        })

    cursos = Curso.objects.filter(activo=True).order_by('nombre')
    anios_disponibles = sorted(
        set(Matricula.objects.dates('fecha_matricula', 'year').values_list('fecha_matricula__year', flat=True)),
        reverse=True
    )

    return render(request, 'historial/lista.html', {
        'estructura': estructura,
        'cursos': cursos,
        'anios': anios_disponibles,
        'meses_es': MESES_ES,
        'filtros': filtros,
        'total_general': qs.count(),
    })


@matricula_requerida
def historial_export(request):
    """
    Descarga del historial como Excel. El archivo tiene una hoja por año
    (o una sola si se filtró por año específico).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    qs, filtros = _filtrar_matriculas(request)
    qs = qs.order_by('-fecha_matricula', '-id')

    # Agrupar por año
    por_anio = defaultdict(list)
    for m in qs:
        por_anio[m.fecha_matricula.year].append(m)

    if not por_anio:
        # Excel vacío con mensaje
        return _build_excel_response(
            filename='historial_vacio.xlsx',
            sheet_name='Historial',
            headers=['Sin datos'],
            rows=[['No hay matrículas con los filtros aplicados.']],
        )

    # Construir el archivo manualmente con varias hojas
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='1A237E')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    month_font = Font(bold=True, color='1A237E', size=12)
    month_fill = PatternFill('solid', fgColor='FFF8E1')
    total_font = Font(bold=True, color='2E7D32', size=10)
    total_fill = PatternFill('solid', fgColor='E8F5E9')
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )

    headers = [
        'Fecha matrícula', 'Cédula', 'Apellidos y Nombres',
        'Curso', 'Modalidad', 'Categoría', 'Sede',
        'Valor curso', 'Pagado', 'Saldo', 'Estado',
    ]

    for anio in sorted(por_anio.keys(), reverse=True):
        ws = wb.create_sheet(title=f'Año {anio}')

        # Título
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title = ws.cell(row=1, column=1, value=f'Historial de matrículas — {anio}')
        title.font = Font(bold=True, size=14, color='1A237E')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 24

        # Encabezados
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        ws.row_dimensions[2].height = 30

        # Agrupar por mes dentro del año
        por_mes = defaultdict(list)
        for m in por_anio[anio]:
            por_mes[m.fecha_matricula.month].append(m)

        current_row = 3
        total_anio_facturado = Decimal('0.00')
        total_anio_pagado = Decimal('0.00')
        total_anio_saldo = Decimal('0.00')

        for mes in sorted(por_mes.keys(), reverse=True):
            # Fila separadora del mes
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=len(headers),
            )
            mes_cell = ws.cell(
                row=current_row, column=1,
                value=f'▸ {MESES_ES[mes]} {anio}  ({len(por_mes[mes])} matrícula(s))'
            )
            mes_cell.font = month_font
            mes_cell.fill = month_fill
            mes_cell.alignment = Alignment(horizontal='left', vertical='center')
            current_row += 1

            mes_facturado = Decimal('0.00')
            mes_pagado = Decimal('0.00')
            mes_saldo = Decimal('0.00')

            for m in por_mes[mes]:
                row_data = [
                    m.fecha_matricula.strftime('%d/%m/%Y'),
                    m.estudiante.cedula,
                    m.estudiante.nombre_completo,
                    m.curso.nombre,
                    m.get_modalidad_display(),
                    m.curso.categoria.nombre if m.curso.categoria else '—',
                    m.sede,
                    float(m.valor_curso or 0),
                    float(m.valor_pagado or 0),
                    float(m.saldo or 0),
                    m.estado_pago,
                ]
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center')
                current_row += 1
                mes_facturado += m.valor_curso or Decimal('0.00')
                mes_pagado += m.valor_pagado or Decimal('0.00')
                mes_saldo += m.saldo or Decimal('0.00')

            # Subtotal del mes
            for col_idx in range(1, 8):
                cell = ws.cell(row=current_row, column=col_idx, value='')
                cell.fill = total_fill
                cell.border = thin_border
            ws.cell(row=current_row, column=7, value='Subtotal mes:').font = total_font
            ws.cell(row=current_row, column=7).alignment = Alignment(horizontal='right', vertical='center')

            for col_idx, val in [(8, float(mes_facturado)), (9, float(mes_pagado)),
                                 (10, float(mes_saldo))]:
                c = ws.cell(row=current_row, column=col_idx, value=val)
                c.font = total_font
                c.fill = total_fill
                c.border = thin_border
                c.alignment = Alignment(vertical='center')
            ws.cell(row=current_row, column=11, value='').fill = total_fill
            current_row += 2  # espacio extra antes del próximo mes

            total_anio_facturado += mes_facturado
            total_anio_pagado += mes_pagado
            total_anio_saldo += mes_saldo

        # Total del año
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col_idx, value='')
            cell.fill = PatternFill('solid', fgColor='1A237E')
        ws.cell(row=current_row, column=7, value=f'TOTAL {anio}:').font = Font(bold=True, color='FFFFFF', size=11)
        ws.cell(row=current_row, column=7).fill = PatternFill('solid', fgColor='1A237E')
        ws.cell(row=current_row, column=7).alignment = Alignment(horizontal='right', vertical='center')

        for col_idx, val in [(8, float(total_anio_facturado)), (9, float(total_anio_pagado)),
                             (10, float(total_anio_saldo))]:
            c = ws.cell(row=current_row, column=col_idx, value=val)
            c.font = Font(bold=True, color='FFFFFF', size=11)
            c.fill = PatternFill('solid', fgColor='1A237E')
            c.alignment = Alignment(vertical='center')

        # Auto-ancho
        for col_idx in range(1, len(headers) + 1):
            max_length = len(headers[col_idx - 1])
            for row_idx in range(3, current_row + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is not None:
                    max_length = max(max_length, len(str(v)))
            ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = min(max_length + 3, 38)

        ws.freeze_panes = 'A3'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    fecha_str = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f'historial_matriculados_{fecha_str}.xlsx'

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ═════════════════════════════════════════════════════════════════
# Estudiantes
# ═════════════════════════════════════════════════════════════════

@matricula_requerida
def estudiantes_lista(request):
    """
    Listado de estudiantes con búsqueda. Cada estudiante muestra el conteo
    de cursos matriculados y un enlace al detalle.
    """
    q = request.GET.get('q', '').strip()
    qs = Estudiante.objects.annotate(num_matriculas=Count('matriculas')).order_by('apellidos', 'nombres')

    if q:
        qs = qs.filter(
            Q(cedula__icontains=q)
            | Q(apellidos__icontains=q)
            | Q(nombres__icontains=q)
            | Q(correo__icontains=q)
            | Q(celular__icontains=q)
        )

    return render(request, 'estudiantes/lista.html', {
        'estudiantes': qs,
        'q': q,
        'total': qs.count(),
    })


@matricula_requerida
def estudiantes_por_curso(request):
    """
    Estudiantes agrupados por curso. Útil cuando se quiere ver la nómina
    completa de un curso específico.
    """
    curso_id = request.GET.get('curso', '').strip()
    modalidad = request.GET.get('modalidad', '').strip()

    cursos_qs = Curso.objects.filter(activo=True).order_by('nombre')

    grupos = []
    for curso in cursos_qs:
        if curso_id and str(curso.id) != curso_id:
            continue
        mat_qs = curso.matriculas.select_related('estudiante', 'jornada').order_by(
            'estudiante__apellidos', 'estudiante__nombres'
        )
        if modalidad in ('presencial', 'online'):
            mat_qs = mat_qs.filter(modalidad=modalidad)
        if mat_qs.exists() or not curso_id:
            grupos.append({
                'curso': curso,
                'matriculas': mat_qs,
                'total': mat_qs.count(),
            })

    # Ocultar cursos sin matriculados (excepto si se filtró por curso)
    if not curso_id:
        grupos = [g for g in grupos if g['total'] > 0]

    return render(request, 'estudiantes/por_curso.html', {
        'grupos': grupos,
        'cursos': cursos_qs,
        'curso_seleccionado': curso_id,
        'modalidad': modalidad,
    })


@matricula_requerida
def estudiante_detalle(request, pk):
    """Detalle de un estudiante con todas sus matrículas."""
    estudiante = get_object_or_404(Estudiante, pk=pk)
    matriculas = estudiante.matriculas.select_related(
        'curso', 'curso__categoria', 'jornada'
    ).order_by('-fecha_matricula')

    # Agrupar por año para el "historial"
    por_anio = defaultdict(list)
    for m in matriculas:
        por_anio[m.fecha_matricula.year].append(m)

    historial = []
    for anio in sorted(por_anio.keys(), reverse=True):
        items = por_anio[anio]
        historial.append({
            'anio': anio,
            'matriculas': items,
            'total_facturado': sum((m.valor_curso or Decimal('0.00')) for m in items),
            'total_pagado': sum((m.valor_pagado or Decimal('0.00')) for m in items),
        })

    return render(request, 'estudiantes/detalle.html', {
        'estudiante': estudiante,
        'matriculas': matriculas,
        'historial': historial,
        'total_matriculas': matriculas.count(),
    })


@matricula_requerida
def estudiantes_export(request):
    """
    Descarga el directorio de estudiantes como Excel.
    Si se pasa ?por_curso=1, genera una hoja por curso.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    por_curso = request.GET.get('por_curso', '') == '1'
    q = request.GET.get('q', '').strip()
    curso_id = request.GET.get('curso', '').strip()
    modalidad = request.GET.get('modalidad', '').strip()

    if por_curso:
        # Una hoja por curso (solo cursos con matriculados)
        wb = Workbook()
        wb.remove(wb.active)

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', fgColor='1A237E')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD'),
        )

        headers = [
            'Cédula', 'Apellidos', 'Nombres', 'Edad',
            'Correo', 'Celular', 'Ciudad', 'Nivel',
            'Modalidad', 'Fecha matrícula', 'Valor', 'Pagado', 'Saldo', 'Estado',
        ]

        cursos_qs = Curso.objects.filter(activo=True).order_by('nombre')
        if curso_id and curso_id.isdigit():
            cursos_qs = cursos_qs.filter(id=int(curso_id))

        hojas_creadas = 0
        for curso in cursos_qs:
            mat_qs = curso.matriculas.select_related('estudiante').order_by(
                'estudiante__apellidos', 'estudiante__nombres'
            )
            if modalidad in ('presencial', 'online'):
                mat_qs = mat_qs.filter(modalidad=modalidad)

            if not mat_qs.exists():
                continue

            # Excel limita el nombre de hoja a 31 chars y prohíbe ciertos caracteres
            nombre_hoja = ''.join(c if c not in '\\/:*?[]' else '_' for c in curso.nombre)[:31]
            ws = wb.create_sheet(title=nombre_hoja)
            hojas_creadas += 1

            # Título
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            t = ws.cell(row=1, column=1, value=f'{curso.nombre} — {mat_qs.count()} estudiante(s)')
            t.font = Font(bold=True, size=14, color='1A237E')
            t.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 24

            # Encabezados
            for col_idx, h in enumerate(headers, start=1):
                c = ws.cell(row=2, column=col_idx, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = header_align
                c.border = thin_border
            ws.row_dimensions[2].height = 30

            for row_idx, m in enumerate(mat_qs, start=3):
                e = m.estudiante
                row_data = [
                    e.cedula, e.apellidos, e.nombres, e.edad or '',
                    e.correo or '', e.celular or '', e.ciudad or '',
                    e.get_nivel_formacion_display() if e.nivel_formacion else '',
                    m.get_modalidad_display(),
                    m.fecha_matricula.strftime('%d/%m/%Y') if m.fecha_matricula else '',
                    float(m.valor_curso or 0),
                    float(m.valor_pagado or 0),
                    float(m.saldo or 0),
                    m.estado_pago,
                ]
                for col_idx, val in enumerate(row_data, start=1):
                    c = ws.cell(row=row_idx, column=col_idx, value=val)
                    c.border = thin_border
                    c.alignment = Alignment(vertical='center')

            # Auto-ancho
            for col_idx in range(1, len(headers) + 1):
                max_length = len(headers[col_idx - 1])
                for row_idx in range(3, mat_qs.count() + 3):
                    v = ws.cell(row=row_idx, column=col_idx).value
                    if v is not None:
                        max_length = max(max_length, len(str(v)))
                ws.column_dimensions[
                    ws.cell(row=2, column=col_idx).column_letter
                ].width = min(max_length + 3, 35)

            ws.freeze_panes = 'A3'

        if hojas_creadas == 0:
            ws = wb.create_sheet(title='Sin datos')
            ws.cell(row=1, column=1, value='No hay estudiantes con los filtros aplicados.')

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        fecha_str = datetime.now().strftime('%Y%m%d_%H%M')
        filename = f'estudiantes_por_curso_{fecha_str}.xlsx'
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    # Modo plano: una sola hoja con todos los estudiantes
    estudiantes_qs = Estudiante.objects.annotate(
        num_matriculas=Count('matriculas')
    ).order_by('apellidos', 'nombres')

    if q:
        estudiantes_qs = estudiantes_qs.filter(
            Q(cedula__icontains=q)
            | Q(apellidos__icontains=q)
            | Q(nombres__icontains=q)
        )

    headers = [
        'Cédula', 'Apellidos', 'Nombres', 'Edad',
        'Correo', 'Celular', 'Ciudad', 'Nivel formación',
        'Título profesional', '# Matrículas', 'Cursos',
    ]

    rows = []
    for e in estudiantes_qs:
        cursos_str = ', '.join(
            sorted({m.curso.nombre for m in e.matriculas.all()})
        )
        rows.append([
            e.cedula, e.apellidos, e.nombres, e.edad or '',
            e.correo or '', e.celular or '', e.ciudad or '',
            e.get_nivel_formacion_display() if e.nivel_formacion else '',
            e.titulo_profesional or '',
            e.num_matriculas,
            cursos_str,
        ])

    fecha_str = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f'estudiantes_{fecha_str}.xlsx'
    return _build_excel_response(
        filename=filename,
        sheet_name='Directorio de Estudiantes',
        headers=headers,
        rows=rows,
    )


@matricula_requerida
def estudiante_export(request, pk):
    """Descarga el historial individual de un estudiante."""
    estudiante = get_object_or_404(Estudiante, pk=pk)
    matriculas = estudiante.matriculas.select_related(
        'curso', 'curso__categoria', 'jornada'
    ).order_by('-fecha_matricula')

    headers = [
        'Año', 'Mes', 'Fecha matrícula', 'Curso', 'Modalidad',
        'Categoría', 'Sede', 'Valor', 'Pagado', 'Saldo', 'Estado',
    ]

    rows = []
    total_facturado = Decimal('0.00')
    total_pagado = Decimal('0.00')

    for m in matriculas:
        rows.append([
            m.fecha_matricula.year,
            MESES_ES[m.fecha_matricula.month],
            m.fecha_matricula.strftime('%d/%m/%Y'),
            m.curso.nombre,
            m.get_modalidad_display(),
            m.curso.categoria.nombre if m.curso.categoria else '—',
            m.sede,
            float(m.valor_curso or 0),
            float(m.valor_pagado or 0),
            float(m.saldo or 0),
            m.estado_pago,
        ])
        total_facturado += m.valor_curso or Decimal('0.00')
        total_pagado += m.valor_pagado or Decimal('0.00')

    totals = {
        7: float(total_facturado),
        8: float(total_pagado),
        9: float(total_facturado - total_pagado),
    }

    fecha_str = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f'estudiante_{estudiante.cedula}_{fecha_str}.xlsx'
    return _build_excel_response(
        filename=filename,
        sheet_name=f'{estudiante.apellidos} {estudiante.nombres}'[:31],
        headers=headers,
        rows=rows,
        totals=totals,
    )


# ═════════════════════════════════════════════════════════════════
# Gestión de abonos por matrícula
# ═════════════════════════════════════════════════════════════════

@matricula_requerida
@require_POST
def matricula_activar_retiro(request, pk):
    """Marca la matrícula como Retiro Voluntario, lo que ajusta su saldo a 0."""
    matricula = get_object_or_404(Matricula, pk=pk)
    if matricula.estado != 'retiro_voluntario':
        matricula.estado = 'retiro_voluntario'
        matricula.save(update_fields=['estado'])
        messages.success(request, 'La matrícula ha sido marcada como Retiro Voluntario. El saldo pendiente ahora es $0.00.')
    return redirect('academia:matricula_abonos', pk=matricula.pk)


@matricula_requerida
def matricula_abonos(request, pk):
    """
    Pantalla central de pagos de una matrícula:
    - Resumen (valor curso, pagado, saldo, estado)
    - Historial de abonos
    - Botón "Registrar abono" (modal)
    """
    matricula = get_object_or_404(
        Matricula.objects.select_related(
            'estudiante', 'curso', 'curso__categoria', 'jornada'
        ),
        pk=pk
    )
    abonos = matricula.abonos.select_related('registrado_por').order_by('-fecha', '-creado')

    # Saldo restante para el modal
    saldo_pendiente = matricula.saldo

    # Distribución por método (para mostrar resumen)
    dist_metodo = defaultdict(lambda: {'count': 0, 'total': Decimal('0.00')})
    for a in abonos:
        dist_metodo[a.get_metodo_display()]['count'] += 1
        dist_metodo[a.get_metodo_display()]['total'] += a.monto

    # Form pre-cargado para el modal (fecha=hoy, monto=saldo)
    form_inicial = AbonoForm(
        initial={
            'fecha': date.today(),
            'monto': saldo_pendiente if saldo_pendiente > 0 else None,
            'metodo': 'efectivo',
        },
        matricula=matricula,
    )

    return render(request, 'pagos/matricula_abonos.html', {
        'matricula': matricula,
        'abonos': abonos,
        'saldo_pendiente': saldo_pendiente,
        'dist_metodo': dict(dist_metodo),
        'form': form_inicial,
        'siguiente_recibo': Abono.generar_numero_recibo(),
    })


@matricula_requerida
@require_POST
def abono_crear(request, matricula_pk):
    """Crear un abono nuevo. Llamado desde el modal."""
    matricula = get_object_or_404(Matricula, pk=matricula_pk)
    form = AbonoForm(request.POST, matricula=matricula)

    if form.is_valid():
        abono = form.save(commit=False)
        abono.matricula = matricula
        abono.registrado_por = request.user
        abono.save()
        messages.success(
            request,
            f'Abono registrado: {abono.numero_recibo} por ${abono.monto}. '
            f'Nuevo saldo: ${matricula.saldo}.'
        )
    else:
        # Recopilar errores legibles (sin __all__ ni nombres internos)
        errores = []
        for field, errs in form.errors.items():
            prefijo = '' if field == '__all__' else f'{form.fields[field].label or field}: '
            for err in errs:
                errores.append(f'{prefijo}{err}')
        messages.error(
            request,
            'No se pudo registrar el abono. ' + ' / '.join(errores)
        )

    return redirect('academia:matricula_abonos', pk=matricula_pk)


@matricula_requerida
def abono_editar(request, matricula_pk, abono_pk):
    """Editar un abono existente."""
    matricula = get_object_or_404(Matricula, pk=matricula_pk)
    abono = get_object_or_404(Abono, pk=abono_pk, matricula=matricula)

    if request.method == 'POST':
        form = AbonoForm(request.POST, instance=abono, matricula=matricula)
        if form.is_valid():
            form.save()
            messages.success(request, f'Abono {abono.numero_recibo} actualizado.')
            return redirect('academia:matricula_abonos', pk=matricula_pk)
    else:
        form = AbonoForm(instance=abono, matricula=matricula)

    return render(request, 'pagos/abono_editar.html', {
        'form': form,
        'abono': abono,
        'matricula': matricula,
    })


@matricula_requerida
@require_POST
def abono_eliminar(request, matricula_pk, abono_pk):
    """Eliminar un abono y recalcular el total."""
    matricula = get_object_or_404(Matricula, pk=matricula_pk)
    abono = get_object_or_404(Abono, pk=abono_pk, matricula=matricula)
    numero = abono.numero_recibo
    monto = abono.monto
    abono.delete()
    messages.success(
        request,
        f'Abono {numero} eliminado (${monto}). Saldo recalculado: ${matricula.saldo}.'
    )
    return redirect('academia:matricula_abonos', pk=matricula_pk)


@matricula_requerida
def abonos_export(request):
    """
    Reporte de abonos en Excel — todos los abonos del periodo,
    con filtros por mes, año, método.
    """
    qs = Abono.objects.select_related(
        'matricula', 'matricula__estudiante', 'matricula__curso',
        'registrado_por',
    ).order_by('-fecha', '-creado')

    anio = request.GET.get('anio', '').strip()
    mes = request.GET.get('mes', '').strip()
    metodo = request.GET.get('metodo', '').strip()

    if anio.isdigit():
        qs = qs.filter(fecha__year=int(anio))
    if mes.isdigit() and 1 <= int(mes) <= 12:
        qs = qs.filter(fecha__month=int(mes))
    if metodo in ('efectivo', 'transferencia', 'tarjeta'):
        qs = qs.filter(metodo=metodo)

    headers = [
        'Nº Recibo', 'Fecha', 'Cédula', 'Estudiante', 'Curso',
        'Modalidad', 'Método', 'Monto', 'Valor curso', 'Saldo restante',
        'Registrado por', 'Observaciones',
    ]

    rows = []
    total_monto = Decimal('0.00')
    total_efectivo = Decimal('0.00')
    total_transf = Decimal('0.00')
    total_tarjeta = Decimal('0.00')

    for a in qs:
        m = a.matricula
        rows.append([
            a.numero_recibo,
            a.fecha.strftime('%d/%m/%Y'),
            m.estudiante.cedula,
            m.estudiante.nombre_completo,
            m.curso.nombre,
            m.get_modalidad_display(),
            a.get_metodo_display(),
            float(a.monto),
            float(m.valor_curso or 0),
            float(m.saldo or 0),
            (a.registrado_por.get_full_name() or a.registrado_por.username) if a.registrado_por else '—',
            a.observaciones or '',
        ])
        total_monto += a.monto
        if a.metodo == 'efectivo':
            total_efectivo += a.monto
        elif a.metodo == 'transferencia':
            total_transf += a.monto
        elif a.metodo == 'tarjeta':
            total_tarjeta += a.monto

    totals = {7: float(total_monto)}

    fecha_str = datetime.now().strftime('%Y%m%d_%H%M')
    sufijo = ''
    if anio:
        sufijo += f'_{anio}'
    if mes:
        sufijo += f'_{mes:0>2}' if not mes.startswith('0') else f'_{mes}'
    if metodo:
        sufijo += f'_{metodo}'
    filename = f'abonos{sufijo}_{fecha_str}.xlsx'

    # Construir el archivo con totales por método al final
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = 'Reporte de Abonos'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='1A237E')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )
    total_font = Font(bold=True, color='1A237E', size=11)
    total_fill = PatternFill('solid', fgColor='FFF8E1')
    method_font = Font(bold=True, color='2E7D32', size=11)
    method_fill = PatternFill('solid', fgColor='E8F5E9')

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    t = ws.cell(row=1, column=1, value='Reporte de Abonos')
    t.font = Font(bold=True, size=14, color='1A237E')
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    # Encabezados
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = thin
    ws.row_dimensions[2].height = 30

    # Datos
    for row_idx, row_data in enumerate(rows, start=3):
        for col_idx, val in enumerate(row_data, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.border = thin
            c.alignment = Alignment(vertical='center')

    # Total general
    total_row = len(rows) + 3
    ws.cell(row=total_row, column=6, value='TOTAL GENERAL:').font = total_font
    ws.cell(row=total_row, column=6).alignment = Alignment(horizontal='right')
    ws.cell(row=total_row, column=6).fill = total_fill
    ws.cell(row=total_row, column=6).border = thin
    c = ws.cell(row=total_row, column=8, value=float(total_monto))
    c.font = total_font
    c.fill = total_fill
    c.border = thin

    # Desglose por método
    metodo_row = total_row + 2
    ws.cell(row=metodo_row, column=1, value='💵 Por método de pago:').font = method_font
    metodo_row += 1
    for label, total in [
        ('Efectivo', total_efectivo),
        ('Transferencia', total_transf),
        ('Tarjeta', total_tarjeta),
    ]:
        ws.cell(row=metodo_row, column=1, value=label).font = method_font
        ws.cell(row=metodo_row, column=1).fill = method_fill
        ws.cell(row=metodo_row, column=1).border = thin
        c = ws.cell(row=metodo_row, column=2, value=float(total))
        c.font = method_font
        c.fill = method_fill
        c.border = thin
        metodo_row += 1

    # Auto-ancho
    for col_idx in range(1, len(headers) + 1):
        max_length = len(headers[col_idx - 1])
        for row_idx in range(3, len(rows) + 3):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is not None:
                max_length = max(max_length, len(str(v)))
        ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = min(max_length + 3, 38)

    ws.freeze_panes = 'A3'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@matricula_requerida
def abono_recibo(request, abono_pk):
    """
    Vista del recibo individual (HTML imprimible).
    Cada abono tiene su comprobante.
    """
    abono = get_object_or_404(
        Abono.objects.select_related(
            'matricula', 'matricula__estudiante', 'matricula__curso',
            'registrado_por',
        ),
        pk=abono_pk
    )
    return render(request, 'pagos/recibo.html', {
        'abono': abono,
        'matricula': abono.matricula,
    })


# ═════════════════════════════════════════════════════════════════
# Pagos por Módulo (control semanal del avance del curso)
# ═════════════════════════════════════════════════════════════════

# Tipos de matrícula que SÍ implican una reserva inicial (los únicos que
# hacen sentido para el control de morosidad por módulo).
TIPOS_CON_RESERVA = ('reserva_abono', 'reserva_modulo_1')


def _construir_matriz_pagos(curso_sel, modalidad='', ciudad='',
                            tipo_matricula='', filtro_modulo_estado=''):
    """
    Construye la matriz estudiantes x modulos para un curso.
    filtro_modulo_estado: cadena con formato "<num>_<estado>",
    por ejemplo "2_Parcial" o "1_Pagado". Filtra por el modulo
    y estado indicados. Cadena vacia = sin filtro.
    """
    n_mod = curso_sel.numero_modulos or 1
    modulos = list(range(1, n_mod + 1))

    qs = Matricula.objects.filter(
        curso=curso_sel
    ).select_related(
        'estudiante', 'jornada', 'registrado_por'
    ).prefetch_related('abonos').exclude(estado='retiro_voluntario')

    if modalidad in ('presencial', 'online'):
        qs = qs.filter(modalidad=modalidad)
    if ciudad:
        qs = qs.filter(jornada__ciudad__iexact=ciudad)
    if tipo_matricula:
        qs = qs.filter(tipo_matricula=tipo_matricula)

    matriculas = []
    for m in qs:
        jornada = m.jornada
        valor_modulo = (
            m.valor_neto / Decimal(n_mod) if n_mod > 0 else Decimal('0.00')
        )

        # Distribución EFECTIVA: incluye reserva → módulo 1, derrama excedente
        # al siguiente, y arrastra la fecha del abono que aterrizó en cada
        # módulo. Toda la lógica vive en el modelo (un solo lugar de verdad).
        desglose = m.desglose_pagos_por_modulo()

        modulos_data = [
            {
                'numero': d['numero'],
                'estado': d['estado'],
                'pagado': d['pagado'],
                'esperado': d['esperado'],
                'fecha_pago': d['fecha_ultimo_pago'],
            }
            for d in desglose
        ]
        # Diccionario plano para los cálculos posteriores que sí lo necesitan
        pagos_efectivos = {d['numero']: d['pagado'] for d in desglose}
            
        modulo_control = next(
            (x for x in modulos_data if x['estado'] != 'Pagado'),
            modulos_data[-1] if modulos_data else None,
        )
        
        # Hoja de recaudacion: SOLO pagos tipo 'por_modulo' asignados
        # explicitamente al modulo de control.
        abonos_modulo = []
        recaudado_hoja = Decimal('0.00')
        if modulo_control:
            todos_abonos = list(m.abonos.filter(cuenta_para_saldo=True))
            abonos_modulo = [
                a for a in todos_abonos
                if a.tipo_pago == 'por_modulo'
                and a.numero_modulo == modulo_control['numero']
            ]
            recaudado_hoja = sum(
                (a.monto for a in abonos_modulo), Decimal('0.00')
            )
            
        metodos = sorted({a.get_metodo_display() for a in abonos_modulo})
        tipos = sorted({a.get_tipo_pago_display() for a in abonos_modulo})
        bancos = sorted({a.get_banco_display() for a in abonos_modulo if a.banco})
        recuperaciones_pendientes = m.recuperaciones_pendientes.filter(pagada=False)
        recuperacion_txt = ', '.join(
            f"Mód. {r.numero_modulo}" for r in recuperaciones_pendientes
        )
        
        # Reserva / abono libre = abonos que NO son por_modulo. Estos no
        # entran a la matriz pero sí suman al saldo total. Se muestran como
        # un texto pequeño "Reservado: $X" debajo del valor pagado.
        reserva_total = sum(
            (a.monto for a in m.abonos.filter(cuenta_para_saldo=True).exclude(tipo_pago='por_modulo')),
            Decimal('0.00'),
        )

        matriculas.append({
            'matricula': m,
            'estudiante': m.estudiante,
            'curso_nombre': m.curso.nombre,
            'modulos_data': modulos_data,
            'valor_modulo_sugerido': valor_modulo,
            'tipo_matricula_codigo': m.tipo_matricula,
            'tipo_matricula_label': m.get_tipo_matricula_display(),
            'reserva_total': reserva_total,
            'jornada_inicio': jornada.fecha_inicio if jornada and jornada.fecha_inicio else None,
            'jornada_dia': jornada.descripcion_legible if jornada else '—',
            'jornada_horario': m.horario,
            'jornada_sede': m.sede,
            'jornada_resumen': (
                f"{jornada.descripcion_legible} · "
                f"{jornada.fecha_inicio.strftime('%d/%m/%Y') if jornada and jornada.fecha_inicio else 'Sin fecha'}"
                f"{' · ' + m.horario if m.horario != '—' else ''}"
                f"{' · ' + m.sede if m.sede != '—' else ''}"
            ) if jornada else '—',
            'modulo_control': modulo_control['numero'] if modulo_control else '—',
            'recaudar_control': (
                max(valor_modulo - recaudado_hoja, Decimal('0.00'))
            ),
            'recaudado_control': recaudado_hoja,
            'forma_pago_control': ', '.join(metodos) if metodos else 'Sin pagar',
            'tipo_pago_control': ', '.join(tipos) if tipos else 'Sin pagar',
            'banco_control': ', '.join(bancos) if bancos else '—',
            'asistencia_control': '—',
            'recuperacion_control': recuperacion_txt or '—',
        })

    # ── Filtro por módulo + estado (e.g. "2_Pagado", "2_Pendiente") ──
    # Bajo la regla visual binaria, "Pagado" agrupa cualquier módulo que
    # haya recibido al menos un pago directo (estado interno Pagado o
    # Parcial). "Pendiente" sigue siendo solo los módulos sin pagos.
    if filtro_modulo_estado:
        partes = filtro_modulo_estado.split('_', 1)
        if len(partes) == 2 and partes[0].isdigit() and partes[1] in ('Pagado', 'Parcial', 'Pendiente'):
            num_filtro = int(partes[0])
            est_filtro = partes[1]
            if est_filtro == 'Pagado':
                estados_match = ('Pagado', 'Parcial')
            else:
                estados_match = (est_filtro,)
            matriculas = [
                x for x in matriculas
                if any(
                    mod['numero'] == num_filtro and mod['estado'] in estados_match
                    for mod in x['modulos_data']
                )
            ]

    # ── Resumen por módulo ──
    resumen_lista = []
    for n in modulos:
        pagados = sum(1 for x in matriculas if x['modulos_data'][n - 1]['estado'] == 'Pagado')
        parciales = sum(1 for x in matriculas if x['modulos_data'][n - 1]['estado'] == 'Parcial')
        pendientes = sum(1 for x in matriculas if x['modulos_data'][n - 1]['estado'] == 'Pendiente')
        recaudado = sum(
            (
                x['modulos_data'][n - 1]['pagado']
                for x in matriculas
                if x['modulos_data'][n - 1]['estado'] == 'Pagado'
            ),
            Decimal('0.00')
        )
        resumen_lista.append({
            'numero': n,
            'pagados': pagados,
            'parciales': parciales,
            'pendientes': pendientes,
            'recaudado': recaudado,
            'total_estudiantes': len(matriculas),
        })

    return matriculas, modulos, resumen_lista


@matricula_requerida
def pagos_por_modulo(request):
    """
    Vista MATRIZ: por cada matrícula del curso filtrado, muestra el estado
    de pago de CADA módulo (Pagado / Parcial / Pendiente / sin pagar).

    Filtros: curso (obligatorio para ver detalle), modalidad, ciudad,
             tipo de matrícula, estado del módulo 1.

    Reglas nuevas (v2):
    - El monto de la "reserva" (abonos sin módulo asignado) se distribuye
      automáticamente al primer módulo. Si el estudiante pagó reserva $20
      y luego $60 directos al Módulo 1 (en un curso de $80 / 2 módulos),
      el Módulo 1 aparece como PAGADO.
    - Filtro "tipo de matrícula" para enfocar reservas (los que SÍ tienen
      pendientes que cobrar mes a mes).
    - Filtro "estado del módulo 1" para ver de un vistazo morosos.
    """
    cursos = Curso.objects.filter(activo=True).order_by('nombre')

    curso_id = request.GET.get('curso', '').strip()
    modalidad = request.GET.get('modalidad', '').strip()
    ciudad = request.GET.get('ciudad', '').strip()
    tipo_matricula = request.GET.get('tipo_matricula', '').strip()
    filtro_modulo_estado = request.GET.get('filtro_modulo_estado', '').strip()

    curso_sel = None
    matriculas = []
    modulos = []
    resumen_por_modulo = []

    if curso_id and curso_id.isdigit():
        try:
            curso_sel = Curso.objects.get(pk=int(curso_id), activo=True)
        except Curso.DoesNotExist:
            curso_sel = None

    if curso_sel:
        matriculas, modulos, resumen_por_modulo = _construir_matriz_pagos(
            curso_sel,
            modalidad=modalidad,
            ciudad=ciudad,
            tipo_matricula=tipo_matricula,
            filtro_modulo_estado=filtro_modulo_estado,
        )

    return render(request, 'pagos/por_modulo.html', {
        'cursos': cursos,
        'curso_sel': curso_sel,
        'modulos': modulos,
        'matriculas_data': matriculas,
        'resumen_por_modulo': resumen_por_modulo,
        'tipos_matricula': [
            ('reserva_abono', 'Reserva / Abono'),
            ('reserva_modulo_1', 'Reserva + Módulo 1'),
            ('programa_completo', 'Programa Completo'),
        ],
        'filtros': {
            'curso': curso_id,
            'modalidad': modalidad,
            'ciudad': ciudad,
            'tipo_matricula': tipo_matricula,
            'filtro_modulo_estado': filtro_modulo_estado,
        },
    })


# ═════════════════════════════════════════════════════════════════
# Clases en Recuperación
# ═════════════════════════════════════════════════════════════════

def _filtrar_recuperaciones(request):
    """Aplica los filtros de la tabla de recuperaciones y devuelve queryset + filtros."""
    estado = request.GET.get('estado', 'pendientes').strip() or 'pendientes'
    curso_id = request.GET.get('curso', '').strip()
    q = request.GET.get('q', '').strip()

    if estado not in ('pendientes', 'pagadas', 'todas'):
        estado = 'pendientes'

    qs = RecuperacionPendiente.objects.select_related(
        'matricula', 'matricula__estudiante', 'matricula__curso',
        'matricula__jornada', 'abono',
    )

    if estado == 'pendientes':
        qs = qs.filter(pagada=False)
    elif estado == 'pagadas':
        qs = qs.filter(pagada=True)

    if curso_id and curso_id.isdigit():
        qs = qs.filter(matricula__curso_id=int(curso_id))

    if q:
        qs = qs.filter(
            Q(matricula__estudiante__cedula__icontains=q)
            | Q(matricula__estudiante__apellidos__icontains=q)
            | Q(matricula__estudiante__nombres__icontains=q)
        )

    return qs.order_by('pagada', '-fecha_marcada', '-creado'), {
        'curso': curso_id,
        'q': q,
        'estado': estado,
    }

@matricula_requerida
def recuperaciones_lista(request):
    """
    Listado central de clases en recuperación.
    Muestra: pendientes (sin cobrar) y resueltas (ya cobradas).
    Cada pendiente trae el saldo previo del estudiante.
    """
    qs, filtros = _filtrar_recuperaciones(request)
    cursos = Curso.objects.filter(activo=True).order_by('nombre')

    # Conteos para tarjetas
    total_pendientes = RecuperacionPendiente.objects.filter(pagada=False).count()
    total_pagadas = RecuperacionPendiente.objects.filter(pagada=True).count()

    return render(request, 'pagos/recuperaciones.html', {
        'recuperaciones': qs,
        'cursos': cursos,
        'estado': filtros['estado'],
        'filtros': filtros,
        'total_pendientes': total_pendientes,
        'total_pagadas': total_pagadas,
    })


@matricula_requerida
def recuperaciones_export_excel(request):
    """Exporta la tabla filtrada de clases en recuperación a Excel."""
    recuperaciones, filtros = _filtrar_recuperaciones(request)

    headers = [
        'Estado', 'Fecha falta', 'Cédula', 'Estudiante', 'Curso',
        'Modalidad', 'Módulo', 'Saldo al marcar', 'Fecha recuperación',
        'Recibo', 'Tipo de pago', 'Monto pagado', 'Método', 'Banco / app',
        'Cuenta para saldo', 'Observaciones', 'Asistencia',
    ]

    rows = []
    total_saldo = Decimal('0.00')
    total_pagado = Decimal('0.00')

    for r in recuperaciones:
        abono = r.abono
        estado_label = 'Pagada' if r.pagada else 'Pendiente'
        estudiante = r.matricula.estudiante
        banco = abono.get_banco_display() if (abono and abono.banco) else '—'
        monto = abono.monto if abono else Decimal('0.00')

        rows.append([
            estado_label,
            r.fecha_marcada.strftime('%d/%m/%Y') if r.fecha_marcada else '',
            estudiante.cedula,
            estudiante.nombre_completo,
            r.matricula.curso.nombre,
            r.matricula.get_modalidad_display(),
            r.numero_modulo,
            float(r.saldo_pendiente_al_marcar or 0),
            r.fecha_recuperacion.strftime('%d/%m/%Y') if r.fecha_recuperacion else '—',
            abono.numero_recibo if abono else '—',
            abono.get_tipo_pago_display() if abono else '—',
            float(monto or 0),
            abono.get_metodo_display() if abono else '—',
            banco,
            'Sí' if (abono and abono.cuenta_para_saldo) else ('No' if abono else '—'),
            r.observaciones or '—',
            '',  # Asistencia: en blanco para firma
        ])
        total_saldo += r.saldo_pendiente_al_marcar or Decimal('0.00')
        total_pagado += monto or Decimal('0.00')

    totals = {
        7: float(total_saldo),
        11: float(total_pagado),
    }
    filename = f'recuperaciones_{filtros["estado"]}_{date.today().strftime("%Y%m%d")}.xlsx'
    return _build_excel_response(
        filename=filename,
        sheet_name='Clases en Recuperación',
        headers=headers,
        rows=rows,
        totals=totals,
    )


@matricula_requerida
def recuperaciones_export_pdf(request):
    """Exporta la tabla filtrada de clases en recuperación a PDF horizontal."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        return HttpResponse(
            'Para exportar a PDF instala reportlab: pip install reportlab',
            status=500, content_type='text/plain; charset=utf-8',
        )

    recuperaciones, filtros = _filtrar_recuperaciones(request)
    recuperaciones = list(recuperaciones)

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=0.9*cm, rightMargin=0.9*cm, topMargin=1.1*cm, bottomMargin=0.9*cm,
        title='Clases en Recuperación',
    )
    styles = getSampleStyleSheet()
    titulo_st = ParagraphStyle(
        'titulo_recuperaciones', parent=styles['Title'],
        textColor=colors.HexColor('#1A237E'), fontSize=15,
        alignment=1, spaceAfter=4,
    )
    sub_st = ParagraphStyle(
        'sub_recuperaciones', parent=styles['Normal'],
        textColor=colors.HexColor('#666666'), fontSize=9,
        alignment=1, spaceAfter=10,
    )

    elementos = [
        Paragraph('Clases en Recuperación', titulo_st),
        Paragraph(
            f'Estado: {filtros["estado"].title()} · Generado el '
            f'{date.today().strftime("%d/%m/%Y")} · {len(recuperaciones)} registro(s)',
            sub_st,
        ),
    ]

    data = [[
        'Estado', 'Fecha', 'Estudiante', 'Cédula', 'Curso',
        'Mód.', 'Saldo', 'Pago recuperación', 'Método', 'Obs.', 'Asistencia',
    ]]
    total_saldo = Decimal('0.00')
    total_pagado = Decimal('0.00')

    for r in recuperaciones:
        abono = r.abono
        estudiante = r.matricula.estudiante
        monto = abono.monto if abono else Decimal('0.00')
        metodo = abono.get_metodo_display() if abono else '—'
        if abono and abono.banco:
            metodo = f'{metodo} · {abono.get_banco_display()}'

        data.append([
            'Pagada' if r.pagada else 'Pendiente',
            r.fecha_marcada.strftime('%d/%m/%Y') if r.fecha_marcada else '',
            estudiante.nombre_completo,
            estudiante.cedula,
            r.matricula.curso.nombre,
            f'M{r.numero_modulo}',
            f'${float(r.saldo_pendiente_al_marcar or 0):.2f}',
            f'{abono.numero_recibo} · ${float(monto or 0):.2f}' if abono else 'Por cobrar',
            metodo,
            (r.observaciones or '—')[:60],
            '',  # Asistencia: vacía para firma a mano
        ])
        total_saldo += r.saldo_pendiente_al_marcar or Decimal('0.00')
        total_pagado += monto or Decimal('0.00')

    data.append([
        '', '', '', '', 'TOTAL', '',
        f'${float(total_saldo):.2f}',
        f'${float(total_pagado):.2f}',
        '', '', '',
    ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A237E')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('FONTSIZE', (0, 1), (-1, -2), 7),
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F7F7F7')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFF8E1')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#1A237E')),
        # Línea para firma en la columna Asistencia (última, solo filas de datos)
        ('LINEBELOW', (-1, 1), (-1, -2), 0.5, colors.HexColor('#888888')),
    ]))
    elementos.append(table)
    elementos.append(Spacer(1, 0.2*cm))
    doc.build(elementos)

    pdf_bytes = buf.getvalue()
    buf.close()
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    filename = f'recuperaciones_{filtros["estado"]}_{date.today().strftime("%Y%m%d")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@matricula_requerida
@transaction.atomic
def recuperacion_marcar(request, matricula_pk):
    """
    Marcar una clase como pendiente de recuperación para una matrícula.
    Guarda automáticamente el saldo pendiente al momento.
    """
    matricula = get_object_or_404(Matricula, pk=matricula_pk)

    if request.method == 'POST':
        form = RecuperacionPendienteForm(request.POST, matricula=matricula)
        if form.is_valid():
            recup = form.save(commit=False)
            recup.matricula = matricula
            recup.saldo_pendiente_al_marcar = matricula.saldo
            recup.save()
            messages.success(
                request,
                f'Clase de Módulo {recup.numero_modulo} marcada para recuperación. '
                f'Saldo arrastrado: ${recup.saldo_pendiente_al_marcar:.2f}.'
            )
            return redirect('academia:matricula_abonos', pk=matricula.pk)
    else:
        form = RecuperacionPendienteForm(
            initial={'fecha_marcada': date.today()},
            matricula=matricula,
        )

    return render(request, 'pagos/recuperacion_marcar.html', {
        'form': form,
        'matricula': matricula,
    })


@matricula_requerida
@transaction.atomic
def recuperacion_cobrar(request, recup_pk):
    """
    Cobra la clase de recuperación: crea un Abono con tipo='recuperacion'
    y deja la marca como pagada=True.
    El usuario decide si el cobro suma al saldo del curso o se cobra aparte.
    """
    recup = get_object_or_404(
        RecuperacionPendiente.objects.select_related(
            'matricula', 'matricula__estudiante', 'matricula__curso',
        ),
        pk=recup_pk,
    )

    if recup.pagada:
        messages.info(request, 'Esta clase de recuperación ya fue cobrada.')
        return redirect('academia:recuperaciones_lista')

    matricula = recup.matricula

    if request.method == 'POST':
        form = AbonoForm(request.POST, matricula=matricula)
        # Forzar tipo_pago='recuperacion' y modulo=el de la recuperación
        post = request.POST.copy()
        post['tipo_pago'] = 'recuperacion'
        post['numero_modulo'] = recup.numero_modulo
        form = AbonoForm(post, matricula=matricula)
        if form.is_valid():
            abono = form.save(commit=False)
            abono.matricula = matricula
            abono.tipo_pago = 'recuperacion'
            abono.numero_modulo = recup.numero_modulo
            abono.registrado_por = request.user
            abono.save()
            # Marcar recuperación como pagada
            recup.pagada = True
            recup.fecha_recuperacion = abono.fecha
            recup.abono = abono
            recup.save()
            messages.success(
                request,
                f'Recuperación cobrada: {abono.numero_recibo} por ${abono.monto}. '
                f'{"(Sumó al saldo del curso)" if abono.cuenta_para_saldo else "(Cobrada aparte, no afecta saldo)"}.'
            )
            return redirect('academia:recuperaciones_lista')
    else:
        form = AbonoForm(
            initial={
                'fecha': date.today(),
                'monto': Decimal('25.00'),
                'tipo_pago': 'recuperacion',
                'numero_modulo': recup.numero_modulo,
                'metodo': 'efectivo',
                'cuenta_para_saldo': True,
            },
            matricula=matricula,
        )

    return render(request, 'pagos/recuperacion_cobrar.html', {
        'form': form,
        'recuperacion': recup,
        'matricula': matricula,
    })


@matricula_requerida
@require_POST
def recuperacion_eliminar(request, recup_pk):
    """Eliminar una recuperación pendiente (no pagada)."""
    recup = get_object_or_404(RecuperacionPendiente, pk=recup_pk)
    if recup.pagada:
        messages.error(request, 'No se puede eliminar una recuperación ya pagada. Eliminar el abono asociado.')
        return redirect('academia:recuperaciones_lista')
    recup.delete()
    messages.success(request, 'Recuperación pendiente eliminada.')
    return redirect('academia:recuperaciones_lista')


# ═════════════════════════════════════════════════════════════════
# Hoja de Recaudación imprimible (formato del PDF de Glenda/Kimberly)
# ═════════════════════════════════════════════════════════════════

DIAS_SEMANA_ES = ['LUNES', 'MARTES', 'MIÉRCOLES', 'JUEVES', 'VIERNES', 'SÁBADO', 'DOMINGO']


@matricula_requerida
def hoja_recaudacion(request):
    """
    Vista imprimible: una hoja por curso para una fecha y ciudad dadas.
    Replica el formato de las hojas físicas (Recaudaciones GYE/QUITO).

    Filtros: fecha (obligatoria), ciudad (opcional), curso (opcional).
    Si no se filtra por curso, genera UNA HOJA POR CADA CURSO con
    matrículas activas en esa fecha/ciudad.
    """
    from datetime import datetime as _dt

    fecha_str = request.GET.get('fecha', '').strip()
    ciudad = request.GET.get('ciudad', '').strip()
    curso_id = request.GET.get('curso', '').strip()
    modalidad = request.GET.get('modalidad', '').strip().lower()
    if modalidad not in ('presencial', 'online'):
        modalidad = ''  # vacío = todas las modalidades

    fecha_obj = None
    if fecha_str:
        try:
            fecha_obj = _dt.strptime(fecha_str, '%Y-%m-%d').date()
        except ValueError:
            fecha_obj = None

    cursos_disponibles = Curso.objects.filter(activo=True).order_by('nombre')

    hojas = []  # lista de dicts: {curso, fecha, ciudad, responsable, items[], totales}

    if fecha_obj:
        # Determinar qué cursos incluir
        cursos_qs = Curso.objects.filter(activo=True).order_by('nombre')
        if curso_id and curso_id.isdigit():
            cursos_qs = cursos_qs.filter(pk=int(curso_id))

        for curso in cursos_qs:
            # Matrículas activas de este curso (no retiradas) en la ciudad indicada
            mat_qs = Matricula.objects.filter(
                curso=curso,
            ).exclude(estado='retiro_voluntario').select_related(
                'estudiante', 'jornada', 'registrado_por',
            )
            if ciudad:
                mat_qs = mat_qs.filter(jornada__ciudad__iexact=ciudad)
            if modalidad:
                mat_qs = mat_qs.filter(jornada__modalidad=modalidad)

            if not mat_qs.exists():
                continue

            items = []
            total_efectivo = Decimal('0.00')
            total_transferencia = Decimal('0.00')
            total_recaudar_esperado = Decimal('0.00')
            total_recaudado = Decimal('0.00')

            for m in mat_qs:
                # Abonos del estudiante registrados ese día (todos los métodos)
                abonos_dia = m.abonos.filter(fecha=fecha_obj)
                pagado_dia = sum((a.monto for a in abonos_dia), Decimal('0.00'))

                # Forma de pago del día (concatenadas si hay varias)
                metodos = sorted({a.get_metodo_display() for a in abonos_dia})
                bancos = sorted({a.get_banco_display() for a in abonos_dia if a.banco})
                forma = ', '.join(metodos) if metodos else '—'
                banco_str = ', '.join(bancos) if bancos else '—'

                # Determinar el módulo "actual" (el más alto que tiene un abono o el siguiente pendiente)
                pagos_mod = m.pagos_por_modulo()
                modulo_actual = max(pagos_mod.keys()) if pagos_mod else 1

                # Recuperaciones pendientes en este módulo
                tiene_recup = m.recuperaciones_pendientes.filter(pagada=False).exists()
                recup_str = '✱ Recuperar' if tiene_recup else ''

                # Suma a totales por método
                for a in abonos_dia:
                    if a.metodo == 'efectivo':
                        total_efectivo += a.monto
                    elif a.metodo in ('transferencia', 'tarjeta'):
                        total_transferencia += a.monto

                total_recaudar_esperado += m.saldo if m.saldo > 0 else Decimal('0.00')
                total_recaudado += pagado_dia

                items.append({
                    'estudiante': m.estudiante,
                    'modulo': modulo_actual,
                    'recaudar': m.saldo if m.saldo > 0 else Decimal('0.00'),
                    'recaudado': pagado_dia,
                    'forma_pago': forma,
                    'banco': banco_str,
                    'asistencia': '—',  # no tenemos campo asistencia, queda manual
                    'recuperacion': recup_str,
                    'talla': m.talla_camiseta or '',
                    'jornada_inicio': m.jornada.fecha_inicio if (m.jornada and m.jornada.fecha_inicio) else None,
                })

            # Responsable: usuario que más matrículas registró en ese curso
            responsables = {}
            for m in mat_qs:
                if m.registrado_por_id:
                    nombre = (
                        f'{m.registrado_por.first_name} {m.registrado_por.last_name}'.strip()
                        or m.registrado_por.username
                    )
                    responsables[nombre] = responsables.get(nombre, 0) + 1
            responsable = max(responsables.items(), key=lambda x: x[1])[0] if responsables else '—'

            dia_semana = DIAS_SEMANA_ES[fecha_obj.weekday()]

            hojas.append({
                'curso': curso,
                'fecha': fecha_obj,
                'dia_semana': dia_semana,
                'ciudad': ciudad or '—',
                'responsable': responsable,
                'items': items,
                'total_efectivo': total_efectivo,
                'total_transferencia': total_transferencia,
                'total_recaudar_esperado': total_recaudar_esperado,
                'total_recaudado': total_recaudado,
            })

    return render(request, 'pagos/hoja_recaudacion.html', {
        'cursos_disponibles': cursos_disponibles,
        'hojas': hojas,
        'filtros': {
            'fecha': fecha_str,
            'ciudad': ciudad,
            'curso': curso_id,
            'modalidad': modalidad,
        },
    })


# ═════════════════════════════════════════════════════════════════
# Exportación de "Pagos por Módulo" (Excel y PDF)
# ═════════════════════════════════════════════════════════════════

def _export_pagos_modulo_filtros(request):
    """
    Resuelve el curso seleccionado y otros filtros para las exportaciones
    de Pagos por Módulo. Devuelve (curso_sel | None, dict_filtros).

    NOTA: el formulario de la pantalla y los enlaces de exportación
    arman el querystring con `filtro_modulo_estado` (mismo nombre que
    espera `_construir_matriz_pagos`). Mantener este nombre alineado.
    """
    curso_id = request.GET.get('curso', '').strip()
    modalidad = request.GET.get('modalidad', '').strip()
    ciudad = request.GET.get('ciudad', '').strip()
    tipo_matricula = request.GET.get('tipo_matricula', '').strip()
    filtro_modulo_estado = request.GET.get('filtro_modulo_estado', '').strip()

    curso_sel = None
    if curso_id and curso_id.isdigit():
        try:
            curso_sel = Curso.objects.get(pk=int(curso_id), activo=True)
        except Curso.DoesNotExist:
            curso_sel = None

    return curso_sel, {
        'modalidad': modalidad,
        'ciudad': ciudad,
        'tipo_matricula': tipo_matricula,
        'filtro_modulo_estado': filtro_modulo_estado,
    }


@matricula_requerida
def pagos_por_modulo_export_excel(request):
    """Exporta la matriz de pagos por módulo del curso filtrado a Excel."""
    curso_sel, filtros = _export_pagos_modulo_filtros(request)
    if not curso_sel:
        messages.error(request, 'Selecciona un curso para exportar la matriz.')
        return redirect('academia:pagos_por_modulo')

    matriculas, modulos, _resumen = _construir_matriz_pagos(
        curso_sel,
        modalidad=filtros['modalidad'],
        ciudad=filtros['ciudad'],
        tipo_matricula=filtros['tipo_matricula'],
        filtro_modulo_estado=filtros['filtro_modulo_estado'],
    )

    # Encabezados base + 1 columna por módulo + Asistencia
    headers = [
        'Cédula', 'Estudiante', 'Curso', 'Jornada', 'Día (inicio jornada)',
        'Tipo matrícula', 'Horario', 'Sede',
        'Valor neto', 'Pagado', 'Saldo',
    ]
    for n in modulos:
        headers.append(f'Mód. {n} (estado / pagado)')
    headers.append('Asistencia')

    rows = []
    total_neto = total_pagado = total_saldo = Decimal('0.00')
    for x in matriculas:
        m = x['matricula']
        estu = x['estudiante']
        j = m.jornada
        if j and j.hora_inicio and j.hora_fin:
            horario_txt = f"{j.hora_inicio.strftime('%H:%M')} – {j.hora_fin.strftime('%H:%M')}"
        else:
            horario_txt = '—'
        fila = [
            estu.cedula,
            f'{estu.apellidos} {estu.nombres}'.strip(),
            x['curso_nombre'],
            x['jornada_dia'],
            x['jornada_inicio'].strftime('%d/%m/%Y') if x['jornada_inicio'] else '—',
            x['tipo_matricula_label'],
            horario_txt,
            (j.ciudad if (j and j.ciudad) else '—'),
            float(m.valor_neto or 0),
            float(m.valor_pagado or 0),
            float(m.saldo or 0),
        ]
        for mod in x['modulos_data']:
            fila.append(f"{mod['estado']} – ${float(mod['pagado']):.2f} / ${float(mod['esperado']):.2f}")
        fila.append('')  # Asistencia: en blanco para firmar
        rows.append(fila)
        total_neto += m.valor_neto or Decimal('0.00')
        total_pagado += m.valor_pagado or Decimal('0.00')
        total_saldo += m.saldo or Decimal('0.00')

    totals = {
        8: round(float(total_neto), 2),
        9: round(float(total_pagado), 2),
        10: round(float(total_saldo), 2),
    }
    filename = f'pagos_modulo_{curso_sel.pk}_{date.today().strftime("%Y%m%d")}.xlsx'
    sheet_name = f'Pagos por Módulo - {curso_sel.nombre}'[:31]
    return _build_excel_response(filename, sheet_name, headers, rows, totals=totals)


@matricula_requerida
def pagos_por_modulo_export_pdf(request):
    """Exporta la matriz de pagos por módulo a un PDF horizontal."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, A3, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
        )
    except ImportError:
        return HttpResponse(
            'Para exportar a PDF instala reportlab: pip install reportlab',
            status=500, content_type='text/plain; charset=utf-8',
        )

    curso_sel, filtros = _export_pagos_modulo_filtros(request)
    if not curso_sel:
        messages.error(request, 'Selecciona un curso para exportar la matriz.')
        return redirect('academia:pagos_por_modulo')

    matriculas, modulos, _resumen = _construir_matriz_pagos(
        curso_sel,
        modalidad=filtros['modalidad'],
        ciudad=filtros['ciudad'],
        tipo_matricula=filtros['tipo_matricula'],
        filtro_modulo_estado=filtros['filtro_modulo_estado'],
    )

    # ── Elegir tamaño de página según número de módulos ──
    n_mod = len(modulos)
    # 9 columnas fijas + n_mod + 1 (asistencia)
    n_cols_total = 9 + n_mod + 1
    if n_cols_total > 13:
        page_size = landscape(A3)  # ancho útil ≈ 41 cm
        page_width_cm = 41.0
    else:
        page_size = landscape(A4)  # ancho útil ≈ 28 cm
        page_width_cm = 28.0

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=page_size,
        leftMargin=0.7*cm, rightMargin=0.7*cm, topMargin=1.0*cm, bottomMargin=0.8*cm,
        title=f'Pagos por Módulo — {curso_sel.nombre}',
    )
    styles = getSampleStyleSheet()
    titulo_st = ParagraphStyle('titulo', parent=styles['Title'],
                               textColor=colors.HexColor('#1A237E'),
                               fontSize=14, alignment=1, spaceAfter=4)
    sub_st = ParagraphStyle('sub', parent=styles['Normal'],
                            textColor=colors.HexColor('#666666'),
                            fontSize=9, alignment=1, spaceAfter=10)
    cell_st = ParagraphStyle(
        'cell', parent=styles['Normal'],
        fontName='Helvetica', fontSize=7, leading=8.5,
    )
    cell_bold_st = ParagraphStyle('cell_b', parent=cell_st, fontName='Helvetica-Bold')
    header_st = ParagraphStyle(
        'h_st', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=8,
        textColor=colors.whitesmoke, alignment=1, leading=9,
    )

    elementos = [
        Paragraph(f'Pagos por Módulo — {curso_sel.nombre}', titulo_st),
        Paragraph(
            f'Formación Técnica y Profesional EC · Generado el '
            f'{date.today().strftime("%d/%m/%Y")} · {len(matriculas)} matrícula(s)',
            sub_st,
        ),
    ]

    headers = [
        'Cédula', 'Estudiante', 'Curso', 'Jornada', 'Día',
        'Tipo matrícula', 'Valor', 'Pagado', 'Saldo',
    ] + [f'Mód. {n}' for n in modulos] + ['Asistencia']

    data = [[Paragraph(h, header_st) for h in headers]]
    total_neto = total_pagado = total_saldo = 0.0
    for x in matriculas:
        m = x['matricula']
        e = x['estudiante']
        fila = [
            Paragraph(e.cedula or '', cell_bold_st),
            Paragraph(f'{e.apellidos} {e.nombres}'.strip(), cell_st),
            Paragraph(x['curso_nombre'] or '', cell_st),
            Paragraph(x['jornada_dia'] or '', cell_st),
            Paragraph(x['jornada_inicio'].strftime('%d/%m/%Y') if x['jornada_inicio'] else '—', cell_st),
            Paragraph(x['tipo_matricula_label'] or '', cell_st),
            Paragraph(f'${float(m.valor_neto or 0):.2f}', cell_st),
            Paragraph(f'<font color="#2e7d32"><b>${float(m.valor_pagado or 0):.2f}</b></font>', cell_st),
            Paragraph(f'<font color="{"#c62828" if (m.saldo or 0) > 0 else "#2e7d32"}"><b>${float(m.saldo or 0):.2f}</b></font>', cell_st),
        ]
        for mod in x['modulos_data']:
            estado = mod['estado']
            simbolo = {'Pagado': '✓', 'Parcial': '◐', 'Pendiente': '○'}.get(estado, '')
            color = {'Pagado': '#2e7d32', 'Parcial': '#f0ad4e', 'Pendiente': '#c62828'}.get(estado, '#000')
            fila.append(Paragraph(
                f'<font color="{color}"><b>{simbolo} ${float(mod["pagado"]):.2f}</b></font>',
                cell_st,
            ))
        fila.append('')  # Asistencia: vacío para firma a mano
        data.append(fila)
        total_neto += float(m.valor_neto or 0)
        total_pagado += float(m.valor_pagado or 0)
        total_saldo += float(m.saldo or 0)

    # Fila de totales
    fila_total = [Paragraph('', cell_st)] * 5 + [Paragraph('<b>TOTAL</b>', cell_bold_st)]
    fila_total += [
        Paragraph(f'<b>${total_neto:.2f}</b>', cell_bold_st),
        Paragraph(f'<b>${total_pagado:.2f}</b>', cell_bold_st),
        Paragraph(f'<b>${total_saldo:.2f}</b>', cell_bold_st),
    ]
    fila_total += [Paragraph('', cell_st)] * len(modulos)
    fila_total.append('')
    data.append(fila_total)

    # ── Anchos de columna explícitos: distribuir el ancho útil ──
    # 9 columnas fijas con anchos predefinidos + módulos + asistencia
    fixed_widths_cm = {
        'cedula': 1.9, 'estudiante': 4.2, 'curso': 3.0, 'jornada': 2.6,
        'dia': 1.7, 'tipo': 2.6, 'valor': 1.7, 'pagado': 1.7, 'saldo': 1.7,
    }
    fixed_total = sum(fixed_widths_cm.values())  # ≈ 21.1 cm
    asistencia_cm = 2.5
    remaining = page_width_cm - fixed_total - asistencia_cm
    mod_width_cm = max(1.4, remaining / max(n_mod, 1)) if n_mod else 0

    col_widths = [
        fixed_widths_cm['cedula']*cm,
        fixed_widths_cm['estudiante']*cm,
        fixed_widths_cm['curso']*cm,
        fixed_widths_cm['jornada']*cm,
        fixed_widths_cm['dia']*cm,
        fixed_widths_cm['tipo']*cm,
        fixed_widths_cm['valor']*cm,
        fixed_widths_cm['pagado']*cm,
        fixed_widths_cm['saldo']*cm,
    ] + [mod_width_cm*cm] * n_mod + [asistencia_cm*cm]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A237E')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN',      (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN',     (0, 0), (-1, 0), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('VALIGN',     (0, 1), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 1), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
        ('GRID',       (0, 0), (-1, -1), 0.3, colors.HexColor('#BBBBBB')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F8F9FB')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFF8E1')),
        # Línea de firma en la columna de Asistencia
        ('LINEBELOW', (-1, 1), (-1, -2), 0.5, colors.HexColor('#888888')),
    ]))
    elementos.append(table)
    doc.build(elementos)

    pdf_bytes = buf.getvalue()
    buf.close()
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    filename = f'pagos_modulo_{curso_sel.pk}_{date.today().strftime("%Y%m%d")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ═════════════════════════════════════════════════════════════════
# Exportación de la "Hoja de Recaudación" (Excel y PDF)
# ═════════════════════════════════════════════════════════════════

def _hojas_recaudacion_data(request):
    """
    Re-construye la data que hoja_recaudacion() entrega al template,
    usando los mismos filtros GET. Devuelve (hojas, filtros).
    """
    from datetime import datetime as _dt

    fecha_str = request.GET.get('fecha', '').strip()
    ciudad = request.GET.get('ciudad', '').strip()
    curso_id = request.GET.get('curso', '').strip()
    modalidad = request.GET.get('modalidad', '').strip().lower()
    if modalidad not in ('presencial', 'online'):
        modalidad = ''

    fecha_obj = None
    if fecha_str:
        try:
            fecha_obj = _dt.strptime(fecha_str, '%Y-%m-%d').date()
        except ValueError:
            fecha_obj = None

    hojas = []
    if not fecha_obj:
        return hojas, {
            'fecha': fecha_str, 'ciudad': ciudad,
            'curso': curso_id, 'modalidad': modalidad,
        }

    cursos_qs = Curso.objects.filter(activo=True).order_by('nombre')
    if curso_id and curso_id.isdigit():
        cursos_qs = cursos_qs.filter(pk=int(curso_id))

    for curso in cursos_qs:
        mat_qs = Matricula.objects.filter(curso=curso).exclude(
            estado='retiro_voluntario'
        ).select_related('estudiante', 'jornada', 'registrado_por')
        if ciudad:
            mat_qs = mat_qs.filter(jornada__ciudad__iexact=ciudad)
        if modalidad:
            mat_qs = mat_qs.filter(jornada__modalidad=modalidad)
        if not mat_qs.exists():
            continue

        items = []
        total_efectivo = Decimal('0.00')
        total_transferencia = Decimal('0.00')
        total_recaudar_esperado = Decimal('0.00')
        total_recaudado = Decimal('0.00')

        for m in mat_qs:
            abonos_dia = m.abonos.filter(fecha=fecha_obj)
            pagado_dia = sum((a.monto for a in abonos_dia), Decimal('0.00'))

            metodos = sorted({a.get_metodo_display() for a in abonos_dia})
            bancos = sorted({a.get_banco_display() for a in abonos_dia if a.banco})
            forma = ', '.join(metodos) if metodos else '—'
            banco_str = ', '.join(bancos) if bancos else '—'

            pagos_mod = m.pagos_por_modulo()
            modulo_actual = max(pagos_mod.keys()) if pagos_mod else 1

            tiene_recup = m.recuperaciones_pendientes.filter(pagada=False).exists()
            recup_str = '✱ Recuperar' if tiene_recup else ''

            for a in abonos_dia:
                if a.metodo == 'efectivo':
                    total_efectivo += a.monto
                elif a.metodo in ('transferencia', 'tarjeta'):
                    total_transferencia += a.monto

            total_recaudar_esperado += m.saldo if m.saldo > 0 else Decimal('0.00')
            total_recaudado += pagado_dia

            items.append({
                'estudiante': m.estudiante,
                'modulo': modulo_actual,
                'recaudar': m.saldo if m.saldo > 0 else Decimal('0.00'),
                'recaudado': pagado_dia,
                'forma_pago': forma,
                'banco': banco_str,
                'asistencia': '—',
                'recuperacion': recup_str,
                'talla': m.talla_camiseta or '',
                'jornada_inicio': m.jornada.fecha_inicio if (m.jornada and m.jornada.fecha_inicio) else None,
                'jornada_descripcion': m.jornada.descripcion_legible if m.jornada else '—',
            })

        responsables = {}
        for m in mat_qs:
            if m.registrado_por_id:
                nombre = (
                    f'{m.registrado_por.first_name} {m.registrado_por.last_name}'.strip()
                    or m.registrado_por.username
                )
                responsables[nombre] = responsables.get(nombre, 0) + 1
        responsable = max(responsables.items(), key=lambda x: x[1])[0] if responsables else '—'

        dia_semana = DIAS_SEMANA_ES[fecha_obj.weekday()]

        hojas.append({
            'curso': curso,
            'fecha': fecha_obj,
            'dia_semana': dia_semana,
            'ciudad': ciudad or '—',
            'responsable': responsable,
            'items': items,
            'total_efectivo': total_efectivo,
            'total_transferencia': total_transferencia,
            'total_recaudar_esperado': total_recaudar_esperado,
            'total_recaudado': total_recaudado,
        })

    return hojas, {'fecha': fecha_str, 'ciudad': ciudad, 'curso': curso_id, 'modalidad': modalidad}


@matricula_requerida
def hoja_recaudacion_export_excel(request):
    """Exporta las hojas de recaudación del día a Excel (todos los cursos en una sola hoja)."""
    hojas, filtros = _hojas_recaudacion_data(request)
    if not hojas:
        messages.error(
            request,
            'No hay hojas para exportar. Verifica que la fecha tenga matrículas.'
        )
        return redirect('academia:hoja_recaudacion')

    headers = [
        'Curso', 'Fecha', 'Día', 'Ciudad', 'Responsable', '#',
        'Estudiante', 'Inicio jornada', 'Mód.',
        'A Recaudar', 'Recaudado', 'Forma de pago', 'Banco', 'Recuperación', 'Talla',
    ]
    rows = []
    total_recaudar = total_recaudado = 0.0
    for h in hojas:
        for idx, item in enumerate(h['items'], start=1):
            rows.append([
                h['curso'].nombre,
                h['fecha'].strftime('%d/%m/%Y'),
                h['dia_semana'],
                h['ciudad'],
                h['responsable'],
                idx,
                item['estudiante'].nombre_completo if hasattr(item['estudiante'], 'nombre_completo')
                else f"{item['estudiante'].apellidos} {item['estudiante'].nombres}".strip(),
                item['jornada_inicio'].strftime('%d/%m/%Y') if item['jornada_inicio'] else '—',
                item['modulo'],
                float(item['recaudar'] or 0),
                float(item['recaudado'] or 0),
                item['forma_pago'],
                item['banco'],
                item['recuperacion'],
                item['talla'],
            ])
            total_recaudar += float(item['recaudar'] or 0)
            total_recaudado += float(item['recaudado'] or 0)

    totals = {
        9: round(total_recaudar, 2),
        10: round(total_recaudado, 2),
    }
    filename = f'hoja_recaudacion_{filtros["fecha"]}.xlsx'
    sheet_name = f'Recaudación {filtros["fecha"]}'[:31]
    return _build_excel_response(filename, sheet_name, headers, rows, totals=totals)


@matricula_requerida
def hoja_recaudacion_export_pdf(request):
    """Exporta las hojas de recaudación del día a un PDF (una página por curso)."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
        )
    except ImportError:
        return HttpResponse(
            'Para exportar a PDF instala reportlab: pip install reportlab',
            status=500, content_type='text/plain; charset=utf-8',
        )

    hojas, filtros = _hojas_recaudacion_data(request)
    if not hojas:
        messages.error(
            request,
            'No hay hojas para exportar. Verifica que la fecha tenga matrículas.'
        )
        return redirect('academia:hoja_recaudacion')

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1*cm, rightMargin=1*cm, topMargin=1.2*cm, bottomMargin=1*cm,
        title=f'Hoja de Recaudación — {filtros["fecha"]}',
    )
    styles = getSampleStyleSheet()
    titulo_st = ParagraphStyle('titulo', parent=styles['Title'],
                               textColor=colors.HexColor('#1A237E'),
                               fontSize=14, alignment=1, spaceAfter=4)
    sub_st = ParagraphStyle('sub', parent=styles['Normal'],
                            textColor=colors.HexColor('#666666'),
                            fontSize=9, alignment=1, spaceAfter=10)
    meta_st = ParagraphStyle('meta', parent=styles['Normal'],
                             fontSize=9, spaceAfter=6)

    elementos = []

    for idx_hoja, h in enumerate(hojas):
        elementos.append(Paragraph(f'Recaudación — {h["curso"].nombre}', titulo_st))
        elementos.append(Paragraph(
            f'<b>Fecha:</b> {h["dia_semana"]} {h["fecha"].strftime("%d/%m/%Y")} · '
            f'<b>Ciudad:</b> {h["ciudad"]} · <b>Responsable:</b> {h["responsable"]} · '
            f'<b>Estudiantes:</b> {len(h["items"])}',
            meta_st,
        ))

        headers = [
            '#', 'Estudiante', 'Inicio jornada', 'Mód.',
            'Recaudar', 'Recaudado', 'Forma', 'Banco', 'Recuperación',
        ]
        data = [headers]
        for i, item in enumerate(h['items'], start=1):
            est = item['estudiante']
            nombre = (est.nombre_completo if hasattr(est, 'nombre_completo')
                      else f'{est.apellidos} {est.nombres}'.strip())
            data.append([
                str(i),
                nombre,
                item['jornada_inicio'].strftime('%d/%m/%Y') if item['jornada_inicio'] else '—',
                str(item['modulo']),
                f"${float(item['recaudar']):.2f}",
                f"${float(item['recaudado']):.2f}",
                item['forma_pago'],
                item['banco'],
                item['recuperacion'],
            ])
        # Fila de totales
        data.append([
            '', 'TOTAL', '', '',
            f"${float(h['total_recaudar_esperado']):.2f}",
            f"${float(h['total_recaudado']):.2f}",
            f"Efectivo: ${float(h['total_efectivo']):.2f}",
            f"Transf.: ${float(h['total_transferencia']):.2f}",
            '',
        ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F0AD4E')),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, 0), 8),
            ('ALIGN',      (0, 0), (-1, 0), 'CENTER'),
            ('FONTSIZE',   (0, 1), (-1, -2), 7),
            ('GRID',       (0, 0), (-1, -1), 0.3, colors.HexColor('#CCCCCC')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#FAFAFA')]),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFF8E1')),
            ('FONTNAME',   (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('TEXTCOLOR',  (0, -1), (-1, -1), colors.HexColor('#1A237E')),
            ('FONTSIZE',   (0, -1), (-1, -1), 8),
        ]))
        elementos.append(table)

        if idx_hoja < len(hojas) - 1:
            elementos.append(PageBreak())

    doc.build(elementos)
    pdf_bytes = buf.getvalue()
    buf.close()
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    filename = f'hoja_recaudacion_{filtros["fecha"]}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ═════════════════════════════════════════════════════════════════
# Alertas de pago pendiente (módulo 1 sin pagar tras inicio de jornada)
# ═════════════════════════════════════════════════════════════════

def _calcular_alertas_pago(usuario_actual=None):
    """
    Devuelve la lista de alertas activas: matrículas con tipo "Reserva/Abono"
    o "Reserva + Módulo 1" cuya jornada YA inició y cuyo módulo 1 sigue
    sin pagar al día siguiente o más tarde.

    Excluye las que ya fueron marcadas como "revisadas hoy".
    """
    from .models import AlertaPagoRevisada
    from datetime import timedelta

    hoy = date.today()

    # Solo revisamos matrículas activas con tipo de matrícula que implica reserva
    qs = Matricula.objects.filter(
        tipo_matricula__in=TIPOS_CON_RESERVA,
        jornada__fecha_inicio__lt=hoy,  # la jornada ya empezó (al menos hace 1 día)
    ).exclude(estado='retiro_voluntario').select_related(
        'estudiante', 'curso', 'jornada'
    ).prefetch_related('abonos')

    # Set de (matricula_id, modulo) ya revisadas hoy → para excluir
    revisadas_hoy = set(
        AlertaPagoRevisada.objects.filter(fecha=hoy)
        .values_list('matricula_id', 'numero_modulo')
    )

    alertas = []
    for m in qs:
        # ¿El módulo 1 sigue pendiente o parcial (no Pagado)?
        # Bajo la regla actual, un módulo solo está "Pagado" si recibió pagos
        # directos (tipo_pago='por_modulo') que cubren su valor. La reserva
        # NO cuenta para esta alerta — si el estudiante solo pagó reserva
        # pero no abonó al módulo 1, sigue atrasado en su módulo 1.
        n_mod = m.curso.numero_modulos or 1
        valor_modulo = (
            m.valor_neto / Decimal(n_mod) if n_mod > 0 else Decimal('0.00')
        )
        desglose_alert = m.desglose_pagos_por_modulo()
        mod1 = desglose_alert[0] if desglose_alert else None
        pagado_m1 = mod1['pagado'] if mod1 else Decimal('0.00')

        if pagado_m1 >= valor_modulo and valor_modulo > 0:
            continue  # módulo 1 pagado → no hay alerta

        if (m.pk, 1) in revisadas_hoy:
            continue  # ya revisada hoy

        dias_atraso = (hoy - m.jornada.fecha_inicio).days
        celular = (m.estudiante.celular or '').strip()
        # Limpieza básica del celular para wa.me (solo dígitos, agregamos 593 si parece local)
        digitos = ''.join(c for c in celular if c.isdigit())
        if digitos.startswith('0') and len(digitos) == 10:
            celular_wa = '593' + digitos[1:]
        elif digitos.startswith('593'):
            celular_wa = digitos
        else:
            celular_wa = digitos  # asumimos que ya viene en formato internacional

        alertas.append({
            'matricula': m,
            'estudiante': m.estudiante,
            'curso': m.curso,
            'jornada': m.jornada,
            'fecha_inicio_jornada': m.jornada.fecha_inicio,
            'dias_atraso': dias_atraso,
            'tipo_matricula_label': m.get_tipo_matricula_display(),
            'pagado_m1': pagado_m1,
            'valor_m1': valor_modulo,
            'saldo_m1': max(valor_modulo - pagado_m1, Decimal('0.00')),
            'saldo_total': m.saldo,
            'celular': celular,
            'celular_wa': celular_wa,
        })

    # Ordenar de mayor atraso a menor
    alertas.sort(key=lambda x: (-x['dias_atraso'], x['matricula'].pk))
    return alertas


@matricula_requerida
@require_POST
def alerta_marcar_revisada(request, matricula_pk):
    """
    Marca como "revisada hoy" la alerta de pago pendiente del módulo 1
    para la matrícula indicada. Esto la oculta del dashboard hasta mañana.
    """
    from .models import AlertaPagoRevisada

    matricula = get_object_or_404(Matricula, pk=matricula_pk)
    numero_modulo = int(request.POST.get('numero_modulo', '1') or 1)
    notas = (request.POST.get('notas', '') or '').strip()

    AlertaPagoRevisada.objects.update_or_create(
        matricula=matricula,
        numero_modulo=numero_modulo,
        fecha=date.today(),
        defaults={
            'revisada_por': request.user,
            'notas': notas,
        },
    )
    messages.success(
        request,
        f'Alerta de {matricula.estudiante} marcada como revisada por hoy. '
        'Si mañana sigue pendiente, volverá a aparecer.'
    )

    redirect_to = request.POST.get('next') or 'academia:bienvenida'
    return redirect(redirect_to)